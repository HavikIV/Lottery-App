'''
Created on May 6, 2015

@author: havik
'''
from kivy.app import App
from kivy.uix.scatter import Scatter
from kivy.uix.label import Label
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.textinput import TextInput
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.gridlayout import GridLayout
from kivy.uix.relativelayout import RelativeLayout
from kivy.core.window import Window
from kivy.uix.image import Image
from kivy.graphics import BorderImage
from kivy.graphics import Color, Rectangle, Line
from kivy.uix.button import Button
from kivy.uix.popup import Popup


import random
#import matplotlib.pyplot as plt
import pandas as pd
#import pylab as pl
import requests
import openpyxl as xl
#from operator import itemgetter
from collections import Counter
from lxml import html
import numpy as np

class HomeScreen(Screen):
    pass

class LotterySelcetionScreen(Screen):
    pass

class TableScreen(Screen):
    pass
        
class TutorialApp(App):
    
    screen_manager = None
    lotteryName = ''
    
    #function to load the table form the excel file corresponding to the passed sheet name
    def loadTable(self, sheetName):
        lotteryData = pd.ExcelFile("Lottery databases.xlsx") #grabs and loads the file into memory
        df = lotteryData.parse(sheetName) #loads the data table form the corresponding sheetName into the df data frame
        return df
         
    #function to display the table
    def showTable(self, table):
        #get the number of rows the table has
        no_of_rows = len(table.index)
        #display the table
         
        return table.head(no_of_rows)
     
    #function to display pie charts of a specific column within the database
    #table is the database that the function will be working with
    #and column is a numberical vaule of which column to get the data from
#     def printPieChart(self, table, column):
#         if column == 6:
#             columnList = table.iloc[:, -1:].values.T.ravel()
#         else:
#             columnList = table.iloc[:, (column - 7): (column - 6)].values.T.ravel()
#         countedList = Counter(columnList)
#           
#         #set up the size of the pie chart
#         fig = plt.figure(figsize=[10, 10])
#         ax = fig.add_subplot(111)
#         cmap = plt.prism()
#           
#         #input variables for pie function
#         slices = [float(v) for v in countedList.values()]
#         colors = cmap(np.linspace(0., 1., len(slices)))
#         labels = [float(k) for k in countedList]
#         columnHeaders = list(table.columns.values)
#           
#         #the pie chart
#         pie_wedge_collection = ax.pie(slices, colors = colors, labels = labels, labeldistance = 1.05, autopct = '%1.1f%%')
#         #get rid of the black outlines between the wedges and around the pie
#         for pie_wedge in pie_wedge_collection[0]:
#             pie_wedge.set_edgecolor('white')
#         ax.set_title(columnHeaders[column + 1])
#         #can't display a Legends as there's too many for plt.legends() too handle
#         #return pyplot.pie([float(v) for v in countedList.values()], labels = [float(k) for k in countedList])
         
    def updateDatabase(self):
        wb = xl.load_workbook("Lottery databases.xlsx") #load the workbook into memory
         
        #variable to hold strings
        popupStrings = [] #empty list
        #list of the sheet names within the workbook
        sheetnames = ["SuperLotto", "MegaMillions", "Powerball"]
        days = ["Fri. ", "Wed. ", "Tue. ", "Sat. "] #days the draws on done on
        #list of the webpages to use grab the new draws
        webPages = ['http://www.calottery.com/play/draw-games/superlotto-plus/winning-numbers', 'http://www.calottery.com/play/draw-games/mega-millions/winning-numbers', 'http://www.calottery.com/play/draw-games/powerball/winning-numbers']
        x = 3
        while x != 0:
            ws = wb.get_sheet_by_name(sheetnames[x-1]) # which sheet to update
            rowIndex = ws.get_highest_row() # gets the highest row index in the sheet
            #lastCellValue = ws.cell(row = rowIndex - 1, column = 0).value #gets the last value in the first column, draw number
            lastCellValue = ws.cell(row = rowIndex, column = 1).value #gets the last value in the first column, draw number, for openpyxl 2.0.0+
            page = requests.get(webPages[x-1]) #grabs the webpage needed
            tree = html.fromstring(page.text) #puts the webpage into a tree structure to make it easy to traverse
            #get the newest draw and date from the webpage for comparasion purposes
            draw_and_date = tree.xpath('//*[@id="objBody_content_0_pagecontent_0_objPastWinningNumbers_rptPast_ctl01_lblDrawDateNumber"]/text()')
            #if the table is up to date, it will move on to the next table else it will update it 
            y = int(draw_and_date[0][-4:]) - int(lastCellValue) # checks to see how many draws are missing from the table
            if y == 0:
                #print("The table for " + sheetnames[x-1] + " is up to date.")
                popupStrings.append("The table for " + sheetnames[x-1] + " is up to date.")
                x -= 1 #decrement x by 1 to move on to the next table
            else:
                #while loop to check if the table needs to be updated or not, if yes it will update it
                while y != 0:
                    #grabs the draw and date of the missing draws from the table
                    draw_and_date = tree.xpath('//*[@id="objBody_content_0_pagecontent_0_objPastWinningNumbers_rptPast_ctl0' + str(y) + '_lblDrawDateNumber"]/text()')
                    numbers = tree.xpath(".//*[@id='content']/div[3]/table/tr[" + str(y) + "]/td[2]/span/text()") #numbers
                    numbers = [int(_) for _ in numbers] # converts the text to integers, seems to the convention to use (_) as a variable for these type of situations
                    numbers.sort() #sort the number from smallest to largest
                    mega = tree.xpath(".//*[@id='content']/div[3]/table/tr[" + str(y) + "]/td[3]/text()") #mega number
                    mega = int(mega[0]) # converts the text to integers
                    #write to the file
                    if sheetnames[x-1] == "MegaMillions":
                        d = 0
                    else:
                        d = 1
                    if int(draw_and_date[0][-4:]) % 2 == 0:
                        # if the draw date is even then the day is a Tuesday/Saturday
                        ws.append([int(draw_and_date[0][-4:]), (days[d+2] + draw_and_date[0][:12]), numbers[0], numbers[1], numbers[2], numbers[3], numbers[4], mega]) # print the draw date
                    else:
                        # if the draw date is odd then the day is a Wednesday/Friday
                        ws.append([int(draw_and_date[0][-4:]), (days[d] + draw_and_date[0][:12]), numbers[0], numbers[1], numbers[2], numbers[3], numbers[4], mega])
                    y -= 1 #decrement y by 1 to get the next missing draw
                #print("Updated the " + sheetnames[x-1] + " table successfully!")
                popupStrings.append("Updated the " + sheetnames[x-1] + " table successfully!")
                x -= 1 #decrement x by 1 to move on to the next table
        wb.save("Lottery databases.xlsx") #save the workbook
        #print("Saved the database Sucessfully!")
        popupMessage = Popup(title = 'Update Databases', content = Label(text = popupStrings[0] + '\n' + popupStrings[1] + '\n' + popupStrings[2]), size_hint=(None,None), size = (350, 150))
        popupMessage.open()
     
    # function to get a list of the occurring numbers in each column
    # 6 for the first number columns or 1 for the mega/powerball column
    def getPopularList(self, table, x):
        popular_list = list()
        if x != 1:
            while x != 0:
                column_list = table.iloc[:, (0 - x): (1 - x)].values.T.ravel() # the all of the values in the column
                counted_list = Counter(column_list) # counts how many time each value occurs within the column
                top_five = counted_list.most_common(5) # the top five within the column
                top_five.sort() # sorts the top five
                popular_list = popular_list + top_five
                x -= 1 # decrement x by 1
        else:
            column_list = table.iloc[:, (0 - x):].values.T.ravel() # the all of the values in the column
            counted_list = Counter(column_list) # counts how many time each value occurs within the column
            top_ten = counted_list.most_common(10) # the top five within the column
            top_ten.sort() # sorts the top five
            popular_list = popular_list + top_ten
        #popular_list is actually a tuple of list which contains the value and how many times that value occured
        #but we only want the values by themselves
        popular_list_values = [y[0] for y in popular_list] #this gives us a list of the values
        return popular_list_values
             
    # Function to generate a ticket based on the popular numbers in the lottery
    def generate_ticket(self, table):
        firstFive = self.getPopularList(table, 6) # get the popular numbers for the first five slots
        mega = self.getPopularList(table, 1) # gets the popular numbers for the mega/powerball slot
        five = self.getNumbers(firstFive, 5) # gets first five numbers for the ticket
        one = self.getNumbers(mega, 1) #gets the mega/powerball for the ticket
        five.append(one)
        return five
         
    #function to get numbers
    def getNumbers(self, numbers_list, x):
        numbers = list() #empty list
        rand_range_list = list(range(0, len(numbers_list))) #list of the numbers to choice from at random
        if x != 1:
            while x != 0:
                y = random.choice(rand_range_list) #pick a number
                #it's not making the list properly at the moment, FIX IT!
                #numbers = numbers + numbers_list[y] #add a number to our list to return
                numbers.append(numbers_list[y]) #append the number to the end of the list
                rand_range_list.remove(y) #remove y from our rand_range_list to prevent any repeats
                x -= 1
        else:
            numbers = random.choice(numbers_list) # add a number to our list to return
        return numbers
     
    # function to print the ticket based on which lottery it's from
    def printTicket(self, ticket, lottery):
        #small bubblesort for the first five numbers
        for x in range(1, 5):
            for y in range(0, 4):
                if ticket[y] > ticket[y + 1]:
                    temp = ticket[y]
                    ticket[y] = ticket[y + 1]
                    ticket[y + 1] = temp
        #print statements depending on the lottery
        #print the ticket without any brackets
        if lottery != "Powerball":
            return str(ticket)[1:-4] + " Mega " + str(ticket[-1])
        else:
            return str(ticket)[1:-4] + " Powerball " + str(ticket[-1])
        
    def build(self):
        #Window.clearcolor = (1,1,1,1)
        self.screen_manager = ScreenManager()
        self.screen_manager.add_widget(HomeScreen(name='home'))
        self.screen_manager.add_widget(LotterySelcetionScreen(name='lottery_selection'))
        self.screen_manager.add_widget(TableScreen(name='table'))
        
        return self.screen_manager 
    
if __name__ == '__main__':
    TutorialApp().run()